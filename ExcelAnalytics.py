import openpyxl


class ExcelAnalyser:
    
    def __init__(self, file_name, sheet):
        self.wb = openpyxl.load_workbook(file_name)
        #Loads sheet into array, sets headers
        self.ws = self.wb[sheet]
        self.headers = {header: n for n, header in enumerate([cell.value for cell in self.ws[1]])}
        self.data = [[cell.value for cell in row] for row in self.ws[2:self.ws.max_row - 1]]
    
    @staticmethod
    def find_type(x):
        if x is None:
            return(type(None))
        else:
            try:
                a = float(x)
                b = int(a)
                if a == b:
                    return int
                else:
                    return float
            except ValueError:
                return str
                
    
    def analyse(self, column):
        
        col_no = self.headers[column]
        values_dict = {}
        for row in self.data:
            cell = row[col_no]
            if cell is not None:
                #Sets cell value to highest possible type
                cell_value = self.find_type(cell)(cell)
            else:
                cell_value = None
            
            #Increments frequency if exists, else creates bucket
            if cell_value in values_dict:
                values_dict[cell_value] += 1
            else:
                values_dict[cell_value] = 1
        
        #Preprocesses 
        type_order = [int, float, str]
        col_type = int
        nulls = False
        for key in values_dict.keys():
            if key is None:
                nulls = True
            else:
                key_type = self.find_type(key)
                if type_order.index(key_type) > type_order.index(col_type):
                    col_type = key_type
                    
        keyset = [x for x in values_dict.keys() if x is not None]
        
        count_keys = (len(values_dict) - int(None in values_dict))
        count_values = (len(self.data) - values_dict.get(None,0))
        key_percentage = round(count_keys/count_values * 100, 2)
        print("Column: %s\tNumeric: %s\tNullable: %s\tKeys: %s\tCount: %s\tPercent: %s%%"%(column + " " * (10 - len(column)), col_type.__name__, values_dict.get(None,0), str(count_keys) +" " * (5 - len(str(count_keys))), count_values, " " * (5 - len(str(key_percentage))) + str(key_percentage)))
        if key_percentage < 10:
#            print(list(values_dict.keys()))
            print("Categorisable")
        return values_dict
    
file_name = "D:/Python/ExcelAutoProcess/titanic3.xlsx"
sheet = "titanic3"
ea = ExcelAnalyser(file_name, sheet)
#keys = ea.analyse("pclass")
#print(keys)
#for key in keys.keys():
#    print(key, isinstance(key, (int, float, complex, type(None))))
for header in ea.headers.keys():
    ea.analyse(header)

    